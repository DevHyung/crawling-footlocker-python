#-*- encoding:utf8 -*-
import os
import time
from selenium import webdriver
from bs4 import BeautifulSoup
from urllib.request import urlretrieve
from openpyxl import load_workbook, Workbook

if __name__=="__main__":
    Title = ['NEW ARRIVALS', "MEN'S", "WOMEN'S", "KIDS'", 'SALE']

    rate = float(input("지금 1달러가 몇 원인지 입력해주세요 : "))

    url = './chromedriver'  # 드라이브가 있는 경로
    driver = webdriver.Chrome(url)
    driver.get("http://www.footlocker.com")

    countPage = 0
    for i in range(0, len(Title)):

        while True:
            answer = input(Title[i] + "의 데이터를 추출하시겠습니까? y/n : ")
            if answer == "y": break
            elif answer == "n": break
            else: continue

        if answer == 'n' : continue
        elif answer == 'y' :
            if not os.path.isdir(str(i + 1).rjust(2, '0') + " " + Title[i]):
                os.mkdir(str(i + 1).rjust(2, '0') + " " + Title[i])

            wb = Workbook()
            ws = wb.active

            ws.cell(row=1, column=1, value="공급사(경로)")  # 공급사(경로)
            ws.cell(row=1, column=2, value="상품명")  # 상품명
            ws.cell(row=1, column=3, value="상품명(관리용)")  # 상품명(관리용)
            ws.cell(row=1, column=4, value="공급사 상품명")  # 공급사 상품명
            ws.cell(row=1, column=5, value="공급가")  # 공급가
            ws.cell(row=1, column=6, value="변동가")  # 변동가
            ws.cell(row=1, column=7, value="공급가변환($->\)")  # 공급가변환($->\)
            ws.cell(row=1, column=8, value="옵션입력")  # 옵션입력
            ws.cell(row=1, column=9, value="등록일")  # 등록일
            ws.cell(row=1, column=10, value="품절")  # 품절
            ws.cell(row=1, column=11, value="기타")  # 기타

            wb.save(str(i + 1).rjust(2, '0') + " " + Title[i] + "/#" + Title[i] + ".xlsx")

            exNum = 2
            imgNum = 2
            if (i == 0) | (i == len(Title) - 1) : # NEW ARRIVALS, SALE
                driver.find_element_by_xpath('//*[@id="navigation-links"]/ul[1]/li[2]/a').click()
                driver.find_element_by_xpath('//*[@id="shop_sub_menu_top"]/ul/li['+ str(i + 1) + ']').click()
                time.sleep(2)

                bs4 = BeautifulSoup(driver.page_source, 'html.parser')
                List = bs4.find('ul', id='group_Brand').findAll('li')

                for l in range(1, len(List) + 1):  # 브랜드
                    if l == 6: continue

                    driver.find_element_by_xpath('//*[@id="navigation-links"]/ul[1]/li[2]/a').click()
                    driver.find_element_by_xpath('//*[@id="shop_sub_menu_top"]/ul/li[' + str(i + 1) + ']').click()

                    driver.find_element_by_xpath('//*[@id="brand_more"]/a').click()
                    time.sleep(2)

                    data = driver.find_element_by_xpath('//*[@id="group_Brand"]/li[' + str(l) + ']/a')
                    route = Title[i] + "-" + data.text[:data.text.find(" (")]

                    while True:
                        answer = input(route + "의 데이터를 추출하시겠습니까? y/n : ")
                        if answer == "y": break
                        elif answer == "n": break
                        else: continue

                    if answer == 'n': continue
                    elif answer == 'y':
                        data.click()

                        while True: # 페이지
                            countPage = countPage + 1
                            if countPage == 2: # 2 : 1, 3 : 2, 4 : 3
                                driver.close()

                                url = './chromedriver'  # 드라이브가 있는 경로
                                driver = webdriver.Chrome(url)
                                driver.get("http://www.footlocker.com")

                                driver.find_element_by_xpath('//*[@id="navigation-links"]/ul[1]/li[2]/a').click()
                                driver.find_element_by_xpath('//*[@id="shop_sub_menu_top"]/ul/li[' + str(i + 1) + ']').click()

                                time.sleep(2)
                                driver.find_element_by_xpath('//*[@id="brand_more"]/a').click()
                                data = driver.find_element_by_xpath('//*[@id="group_Brand"]/li[' + str(l) + ']/a').click()

                                for next in range(1, countPage):
                                    try: driver.find_element_by_xpath('//*[@id="endecaResultsWrapper"]/div[4]/div/div[3]/a[5]').click()
                                    except:
                                        try: driver.find_element_by_xpath('//*[@id="endecaResultsWrapper"]/div[4]/div/div[3]/a[3]').click()
                                        except: break
                                countPage = 0

                            bs4 = BeautifulSoup(driver.page_source, "html.parser")
                            List = bs4.find('div', id='endeca_search_results').findAll('li')

                            for p in range(0, len(List)): # 상품
                                if (p != 0) & ((p + 1) % 5 == 0):
                                    continue
                                else:
                                    driver.find_element_by_xpath('//*[@id="endeca_search_results"]/ul/li[' + str(p + 1) + ']/span/a[2]/img').click()

                                    # 상품명
                                    elem1 = driver.find_element_by_xpath('//*[@id="product_title"]').text

                                    # 가격
                                    elem2 = driver.find_element_by_xpath('//*[@id="list_price"]').text
                                    elem2 = elem2[8:]
                                    rateElem = float(elem2) * rate

                                    #변동가
                                    try:
                                        changeP = driver.find_element_by_xpath('//*[@id="sale_price"]').text
                                        changeP = changeP[7:]
                                        rateElem = float(changeP) * rate
                                    except:
                                        changeP=''

                                    time.sleep(5)
                                    bs4 = BeautifulSoup(driver.page_source, 'html.parser')
                                    List = bs4.find('span', id='size_selection_list').findAll('a')

                                    # 사이즈
                                    elem3 = '사이즈{'
                                    for o in range(0, len(List)):
                                        elem3 = elem3 + List[o].text + "|"
                                    elem3 = elem3[:len(elem3) - 1] + "}"

                                    # 기타
                                    elem4 = driver.find_element_by_xpath('//*[@id="product_details"]/div[1]').text
                                    elem4 = elem4[11:]

                                    # wb = load_workbook('data.xlsx')
                                    ws = wb.active

                                    ws.cell(row=exNum, column=1, value=route)  # 공급사(경로)
                                    ws.cell(row=exNum, column=2, value=elem1)  # 상품명
                                    ws.cell(row=exNum, column=5, value=elem2)  # 공급가
                                    ws.cell(row=exNum, column=6, value=changeP)  # 변동가
                                    ws.cell(row=exNum, column=7, value=rateElem)  # 공급가변환($->\)
                                    ws.cell(row=exNum, column=8, value=elem3)  # 옵션입력
                                    ws.cell(row=exNum, column=11, value=elem4)  # 기타

                                    wb.save(str(i + 1).rjust(2, '0') + " " + Title[i] + "/#" + Title[i] + ".xlsx")

                                    try:
                                        List = bs4.find('div', id='s7ZoomViewerViewer_swatches').findAll('div', class_='s7thumbcell')
                                    except:
                                        List = bs4.find('div', id='s7MixedMediaViewer_colorSwatches').findAll('div', class_='s7thumbcell')

                                    for u in range(0, len(List)):
                                        try:
                                            url = List[u].find('div', class_='s7thumb')['style']
                                            url = url[url.find("https://") : len(url) - 3]
                                            urlretrieve(url, str(i + 1).rjust(2, '0') + " " + Title[i] + "/" + str(imgNum).rjust(5, '0') + "_img" + str(u + 1).rjust(2, '0') + ".jpg")
                                        except: continue

                                    exNum = exNum + 1
                                    imgNum = imgNum + 1
                                    driver.back()

                            try: driver.find_element_by_xpath('//*[@id="endecaResultsWrapper"]/div[4]/div/div[3]/a[5]').click()
                            except:
                                try: driver.find_element_by_xpath('//*[@id="endecaResultsWrapper"]/div[4]/div/div[3]/a[3]').click()
                                except:
                                    if countPage == 1: countPage = 0
                                    break



            else:  # MEN'S, WOMEN'S, KIDS'
                for j in range(1, 3): # SHOES, CLOTHING
                    time.sleep(2)

                    if j == 1:
                        driver.find_element_by_xpath('//*[@id="navigation-links"]/ul[1]/li[2]/a').click()
                        driver.find_element_by_xpath('//*[@id="shop_sub_menu_top"]/ul/li[' + str(i + 1) + ']').click()

                    time.sleep(2)
                    data = driver.find_element_by_xpath('//*[@id="navigation-dropdown"]/div[1]/div[2]/div[' + str(i + 1) + ']/div[1]/ul[' + str(j) + ']/div/a')
                    route = Title[i] + "-" + data.text

                    while True:
                        answer = input(route + "의 데이터를 추출하시겠습니까? y/n : ")
                        if answer == "y": break
                        elif answer == "n": break
                        else: continue

                    if answer == 'n': continue
                    elif answer == 'y':
                        data.click()

                        bs4 = BeautifulSoup(driver.page_source, 'html.parser')
                        List = bs4.find('ul', id='group_Brand').findAll('li')

                        for l in range(0, len(List) + 1): # 브랜드
                            if l == 6 : continue

                            time.sleep(2)
                            driver.find_element_by_xpath('//*[@id="navigation-links"]/ul[1]/li[2]/a').click()
                            driver.find_element_by_xpath('//*[@id="shop_sub_menu_top"]/ul/li[' + str(i + 1) + ']').click()

                            time.sleep(2)
                            data = driver.find_element_by_xpath('//*[@id="navigation-dropdown"]/div[1]/div[2]/div[' + str(i + 1) + ']/div[1]/ul[' + str(j) + ']/div/a')
                            title = data.text
                            data.click()

                            driver.find_element_by_xpath('//*[@id="brand_more"]/a').click()
                            data = driver.find_element_by_xpath('//*[@id="group_Brand"]/li[' + str(l) + ']/a')
                            route = Title[i] + "-" + title + "-" + data.text[:data.text.find(" (")]

                            while True:
                                answer = input(route + "의 데이터를 추출하시겠습니까? y/n : ")
                                if answer == "y": break
                                elif answer == "n": break
                                else: continue

                            if answer == 'n': continue
                            elif answer == 'y':
                                data.click()

                                while True: # 페이지
                                    countPage = countPage + 1
                                    if countPage == 2: # 2 : 1, 3 : 2, 4 : 3
                                        driver.close()

                                        url = './chromedriver'  # 드라이브가 있는 경로
                                        driver = webdriver.Chrome(url)
                                        driver.get("http://www.footlocker.com")

                                        driver.find_element_by_xpath('//*[@id="navigation-links"]/ul[1]/li[2]/a').click()
                                        driver.find_element_by_xpath('//*[@id="shop_sub_menu_top"]/ul/li[' + str(i + 1) + ']').click()
                                        time.sleep(2)

                                        data = driver.find_element_by_xpath('//*[@id="navigation-dropdown"]/div[1]/div[2]/div[' + str(i + 1) + ']/div[1]/ul[' + str(j) + ']/div/a').click()
                                        driver.find_element_by_xpath('//*[@id="brand_more"]/a').click()
                                        data = driver.find_element_by_xpath('//*[@id="group_Brand"]/li[' + str(l) + ']/a').click()

                                        for next in range(1, countPage):
                                            try: driver.find_element_by_xpath('//*[@id="endecaResultsWrapper"]/div[4]/div/div[3]/a[5]').click()
                                            except:
                                                try: driver.find_element_by_xpath('//*[@id="endecaResultsWrapper"]/div[4]/div/div[3]/a[3]').click()
                                                except: break
                                        countPage = 0

                                    bs4 = BeautifulSoup(driver.page_source, "html.parser")
                                    List = bs4.find('div', id='endeca_search_results').findAll('li')

                                    for p in range(0, len(List)): # 상품
                                        if (p != 0) & ((p + 1) % 5 == 0):
                                            continue
                                        else:
                                            driver.find_element_by_xpath('//*[@id="endeca_search_results"]/ul/li[' + str(p + 1) + ']/span/a[2]/img').click()

                                            # 상품명
                                            elem1 = driver.find_element_by_xpath('//*[@id="product_title"]').text

                                            # 가격
                                            elem2 = driver.find_element_by_xpath('//*[@id="list_price"]').text
                                            elem2 = elem2[8:]
                                            rateElem = float(elem2) * rate

                                            #변동가
                                            try:
                                                changeP = driver.find_element_by_xpath('//*[@id="sale_price"]').text
                                                changeP = changeP[7:]
                                                rateElem = float(changeP) * rate
                                            except:
                                                changeP=''

                                            time.sleep(4)
                                            bs4 = BeautifulSoup(driver.page_source, 'html.parser')
                                            List = bs4.find('span', id='size_selection_list').findAll('a')

                                            # 사이즈
                                            elem3 = '사이즈{'
                                            for o in range(0, len(List)):
                                                elem3 = elem3 + List[o].text + "|"
                                            elem3 = elem3[:len(elem3) - 1] + "}"

                                            # 기타
                                            elem4 = driver.find_element_by_xpath('//*[@id="product_details"]/div[1]').text
                                            elem4 = elem4[11:]

                                            # wb = load_workbook('data.xlsx')
                                            ws = wb.active

                                            ws.cell(row=exNum, column=1, value=route)  # 공급사(경로)
                                            ws.cell(row=exNum, column=2, value=elem1)  # 상품명
                                            ws.cell(row=exNum, column=5, value=elem2)  # 공급가
                                            ws.cell(row=exNum, column=6, value=changeP)  # 변동가
                                            ws.cell(row=exNum, column=7, value=rateElem)  # 공급가변환($->\)
                                            ws.cell(row=exNum, column=8, value=elem3)  # 옵션입력
                                            ws.cell(row=exNum, column=11, value=elem4)  # 기타

                                            wb.save(str(i + 1).rjust(2, '0') + " " + Title[i] + "/#" + Title[i] + ".xlsx")

                                            try:
                                                List = bs4.find('div', id='s7ZoomViewerViewer_swatches').findAll('div', class_='s7thumbcell')
                                            except:
                                                List = bs4.find('div', id='s7MixedMediaViewer_colorSwatches').findAll('div', class_='s7thumbcell')

                                            for u in range(0, len(List)):
                                                try:
                                                    url = List[u].find('div', class_='s7thumb')['style']
                                                    url = url[url.find("https://") : len(url) - 3]
                                                    urlretrieve(url, str(i + 1).rjust(2, '0') + " " + Title[i] + "/" + str(imgNum).rjust(5, '0') + "_img" + str(u + 1).rjust(2, '0') + ".jpg")
                                                except: continue

                                            exNum = exNum + 1
                                            imgNum = imgNum + 1
                                            driver.back()

                                    try: driver.find_element_by_xpath('//*[@id="endecaResultsWrapper"]/div[4]/div/div[3]/a[5]').click()
                                    except:
                                        try: driver.find_element_by_xpath('//*[@id="endecaResultsWrapper"]/div[4]/div/div[3]/a[3]').click()
                                        except:
                                            if countPage == 1: countPage = 0
                                            break